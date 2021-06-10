import openpyxl

def get_quantidade_linha(arquivo, nomePagina):
    workbook = openpyxl.load_workbook(arquivo)
    pagina = workbook.get_sheet_by_name(nomePagina)
    return (pagina.max_row)

def get_quantidade_coluna(arquivo, nomePagina):
    workbook = openpyxl.load_workbook(arquivo)
    pagina = workbook.get_sheet_by_name(nomePagina)
    return (pagina.max_column)

def ler_dados(arquivo, nomePagina, rownum, columnno):
    workbook = openpyxl.load_workbook(arquivo)
    pagina = workbook.get_sheet_by_name(nomePagina)
    return pagina.cell(row=rownum, column=columnno).value