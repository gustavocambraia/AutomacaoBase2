import openpyxl

def get_quantidade_linha(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def get_quantidade_coluna(arquivo, nomePagina):
    workbook = openpyxl.load_workbook(arquivo)
    pagina = workbook.get_sheet_by_name(nomePagina)
    return (pagina.max_column)

def ler_dados(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value